"""
Admin Dashboard – Store-wise Reports API
=========================================
Endpoints:
  GET /api/superadmin/reports/store-wise/summary/
      → per-store: total_revenue, total_orders, avg_order_value
        filterable by period=week|month|year  (+ start_date/end_date for custom)
        export=excel supported

  GET /api/superadmin/reports/store-wise/orders/
      → detailed order list scoped to a store
        period=week|month|year  + optional store_id

  GET /api/superadmin/reports/store-wise/credits/
      → credit-note list/summary scoped to a store
        period=week|month|year  + optional store_id

  GET /api/superadmin/reports/store-wise/revenue/
      → time-series revenue per store
        period=week|month|year  + group_by=day|week|month

  GET /api/superadmin/reports/store-wise/retailer-activity/
      → per-retailer order count, spend, last order date
        period=week|month|year  + optional store_id

All endpoints accept:
  period     : week | month | year | custom   (default: month)
  start_date : YYYY-MM-DD   (used when period=custom)
  end_date   : YYYY-MM-DD   (used when period=custom)
  store_id   : int           (optional — filter single store)
  export     : excel         (returns .xlsx download)
"""

from __future__ import annotations

import logging
from datetime import timedelta, date
from io import BytesIO
from decimal import Decimal

import openpyxl
from django.db.models import Sum, Count, Avg, Max, Q, F, DecimalField
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_period(request) -> tuple[date, date]:
    """
    Resolve start_date / end_date from query params.
    period=week|month|year|custom
    """
    period = request.query_params.get("period", "month").lower()
    today = timezone.now().date()

    if period == "week":
        return today - timedelta(days=6), today
    if period == "year":
        return today.replace(month=1, day=1), today
    if period == "custom":
        sd = parse_date(request.query_params.get("start_date", ""))
        ed = parse_date(request.query_params.get("end_date", ""))
        if sd and ed:
            return sd, ed
    # default → month (MTD)
    return today.replace(day=1), today


def _trunc_fn(group_by: str):
    return {
        "day":   TruncDay,
        "week":  TruncWeek,
        "year":  TruncYear,
    }.get(group_by, TruncMonth)


def _excel_response(data: list[dict], filename: str) -> HttpResponse:
    """Return an xlsx HttpResponse from a list of flat dicts."""
    if not data:
        return Response({"error": "No data to export"}, status=status.HTTP_400_BAD_REQUEST)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename[:30]
    headers = list(data[0].keys())
    ws.append(headers)
    for row in data:
        ws.append([str(v) if v is not None else "" for v in row.values()])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return resp


def _pct_change(curr, prev):
    """% change; returns 0.0 when prev is 0."""
    curr, prev = float(curr or 0), float(prev or 0)
    if prev == 0:
        return 100.0 if curr > 0 else 0.0
    return round((curr - prev) / prev * 100, 1)


def _prev_period_dates(start: date, end: date) -> tuple[date, date]:
    """Return the equal-length period immediately preceding [start, end]."""
    delta = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=delta - 1)
    return prev_start, prev_end


# ─────────────────────────────────────────────────────────────────────────────
# Mixin: SUPERADMIN guard
# ─────────────────────────────────────────────────────────────────────────────

class SuperAdminMixin:
    permission_classes = [IsAuthenticated]

    def _check(self, request) -> Response | None:
        if getattr(request.user, "role", None) != "SUPERADMIN":
            return Response(
                {"error": "Only Super Admin can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )
        return None


# ─────────────────────────────────────────────────────────────────────────────
# 1. Store-wise Summary  (revenue, orders, avg_order_value)
# ─────────────────────────────────────────────────────────────────────────────

class StoreWiseSummaryReportView(SuperAdminMixin, APIView):
    """
    GET /api/superadmin/reports/store-wise/summary/

    Returns per-store KPIs:
      total_revenue, total_orders, avg_order_value
      + % change vs previous equivalent period
      + credit total (APPROVED credit notes)

    Query params:
      period    : week | month | year | custom  (default: month)
      start_date, end_date  – when period=custom
      store_id  : int  – filter a single store
      export    : excel
    """

    def get(self, request):
        err = self._check(request)
        if err:
            return err

        from dreamspharmaapp.models import Store, SalesOrder, CreditNote

        start, end = _resolve_period(request)
        prev_start, prev_end = _prev_period_dates(start, end)
        store_filter = request.query_params.get("store_id")

        stores_qs = Store.objects.filter(is_active=True)
        if store_filter:
            stores_qs = stores_qs.filter(id=store_filter)

        rows = []

        for store in stores_qs:
            def _orders(s, e):
                return SalesOrder.objects.filter(
                    fulfilling_store=store,
                    ord_date__gte=s,
                    ord_date__lte=e,
                )

            curr_orders = _orders(start, end)
            prev_orders = _orders(prev_start, prev_end)

            curr_count = curr_orders.count()
            prev_count = prev_orders.count()

            curr_rev = float(curr_orders.aggregate(t=Sum("order_total"))["t"] or 0)
            prev_rev = float(prev_orders.aggregate(t=Sum("order_total"))["t"] or 0)

            curr_avg = round(curr_rev / curr_count, 2) if curr_count else 0.0
            prev_avg = round(prev_rev / prev_count, 2) if prev_count else 0.0

            # Approved credit notes = credits issued to retailer
            curr_credits = float(
                CreditNote.objects.filter(
                    store=store,
                    status="APPROVED",
                    created_at__date__gte=start,
                    created_at__date__lte=end,
                ).aggregate(t=Sum("amount"))["t"] or 0
            )

            rows.append({
                "store_id":                store.id,
                "store_name":              store.name,
                "city":                    store.city,
                "erp_store_id":            store.store_id,
                # ── Current period ──
                "total_orders":            curr_count,
                "total_revenue":           round(curr_rev, 2),
                "avg_order_value":         curr_avg,
                "total_credits_issued":    round(curr_credits, 2),
                # ── vs. previous period ──
                "orders_change_pct":       _pct_change(curr_count, prev_count),
                "revenue_change_pct":      _pct_change(curr_rev, prev_rev),
                "avg_order_change_pct":    _pct_change(curr_avg, prev_avg),
            })

        # Grand totals
        grand_revenue = sum(r["total_revenue"] for r in rows)
        grand_orders  = sum(r["total_orders"]  for r in rows)
        grand_avg     = round(grand_revenue / grand_orders, 2) if grand_orders else 0.0
        grand_credits = sum(r["total_credits_issued"] for r in rows)

        if request.query_params.get("export") == "excel":
            flat = []
            for r in rows:
                flat.append({
                    "Store": r["store_name"],
                    "City": r["city"],
                    "ERP Store ID": r["erp_store_id"],
                    "Total Orders": r["total_orders"],
                    "Total Revenue (₹)": r["total_revenue"],
                    "Avg Order Value (₹)": r["avg_order_value"],
                    "Credits Issued (₹)": r["total_credits_issued"],
                    "Orders Change %": r["orders_change_pct"],
                    "Revenue Change %": r["revenue_change_pct"],
                })
            return _excel_response(flat, "store_wise_summary")

        return Response({
            "success": True,
            "period": {"start_date": str(start), "end_date": str(end)},
            "summary": {
                "grand_total_revenue": grand_revenue,
                "grand_total_orders":  grand_orders,
                "grand_avg_order_value": grand_avg,
                "grand_credits_issued":  grand_credits,
                "total_stores": len(rows),
            },
            "stores": rows,
        })


# ─────────────────────────────────────────────────────────────────────────────
# 2. Store-wise Order Report
# ─────────────────────────────────────────────────────────────────────────────

class StoreWiseOrderReportView(SuperAdminMixin, APIView):
    """
    GET /api/superadmin/reports/store-wise/orders/

    Always returns an Excel (.xlsx) download.

    Query params:
      period     : week | month | year | custom  (default: month)
      start_date, end_date – when period=custom
      store_id   : int  (optional)
      search     : str  (order_id / retailer name)
    """

    def get(self, request):
        err = self._check(request)
        if err:
            return err

        from dreamspharmaapp.models import SalesOrder

        start, end = _resolve_period(request)
        store_filter = request.query_params.get("store_id")
        search = request.query_params.get("search", "").strip()

        orders = SalesOrder.objects.select_related("fulfilling_store").filter(
            ord_date__gte=start,
            ord_date__lte=end,
        ).order_by("-ord_date")

        if store_filter:
            orders = orders.filter(fulfilling_store_id=store_filter)

        if search:
            orders = orders.filter(
                Q(order_id__icontains=search) |
                Q(cust_name__icontains=search) |
                Q(patient_name__icontains=search)
            )

        rows = []
        for o in orders:
            rows.append({
                "Order ID":   o.order_id,
                "Store":      o.fulfilling_store.name if o.fulfilling_store else "",
                "City":       o.fulfilling_store.city if o.fulfilling_store else "",
                "Retailer":   o.cust_name or o.patient_name or "",
                "Date":       str(o.ord_date) if o.ord_date else "",
                "Items":      o.items.count(),
                "Total (INR)": float(o.order_total),
                "Status": (
                    "Delivered"  if o.dc_conversion_flag else
                    "Dispatched" if o.invoices.exists()  else
                    "Confirmed"  if o.ord_conversion_flag else
                    "Pending"
                ),
                "ERP Ref": o.document_pk or "",
            })

        filename = f"order_report_{start}_to_{end}"
        return _excel_response(rows, filename)


# ─────────────────────────────────────────────────────────────────────────────
# 3. Store-wise Credit Report
# ─────────────────────────────────────────────────────────────────────────────

class StoreWiseCreditReportView(SuperAdminMixin, APIView):
    """
    GET /api/superadmin/reports/store-wise/credits/

    Always returns an Excel (.xlsx) download.
    Includes a per-store summary sheet + a detailed credit note sheet.

    Query params:
      period     : week | month | year | custom  (default: month)
      start_date, end_date – when period=custom
      store_id   : int  (optional)
    """

    def get(self, request):
        err = self._check(request)
        if err:
            return err

        from dreamspharmaapp.models import CreditNote

        start, end = _resolve_period(request)
        store_filter = request.query_params.get("store_id")

        cn_qs = CreditNote.objects.select_related("store", "retailer").filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
        ).order_by("-created_at")

        if store_filter:
            cn_qs = cn_qs.filter(store_id=store_filter)

        # ── Build per-store summary ──────────────────────────────────────
        store_summary: dict[int, dict] = {}
        for row in (
            cn_qs.values("store__id", "store__name", "store__city", "status")
            .annotate(count=Count("credit_note_id"), amount=Sum("amount"))
            .order_by("store__name")
        ):
            sid = row["store__id"] or 0
            if sid not in store_summary:
                store_summary[sid] = {
                    "store_id": sid,
                    "store_name": row["store__name"] or "",
                    "city": row["store__city"] or "",
                    "total_count": 0, "total_amount": 0.0,
                    "approved_count": 0, "approved_amount": 0.0,
                    "pending_count": 0, "pending_amount": 0.0,
                    "rejected_count": 0,
                }
            s = store_summary[sid]
            s["total_count"]  += row["count"]
            s["total_amount"] += float(row["amount"] or 0)
            if row["status"] == "APPROVED":
                s["approved_count"]  += row["count"]
                s["approved_amount"] += float(row["amount"] or 0)
            elif row["status"] == "PENDING":
                s["pending_count"]  += row["count"]
                s["pending_amount"] += float(row["amount"] or 0)
            elif row["status"] == "REJECTED":
                s["rejected_count"] += row["count"]

        # ── Build detail rows ─────────────────────────────────────────────
        detail_rows = []
        for cn in cn_qs:
            detail_rows.append({
                "Credit Note ID": cn.credit_note_id,
                "Store":         cn.store.name if cn.store else "",
                "City":          cn.store.city if cn.store else "",
                "Retailer":      cn.retailer.get_full_name() or cn.retailer.username,
                "Product":       cn.product_name,
                "Qty":           cn.quantity_to_return,
                "Amount (INR)":  float(cn.amount),
                "Reason":        cn.get_reason_display(),
                "Status":        cn.status,
                "Date":          str(cn.created_at.date()),
            })

        # ── Two-sheet workbook ─────────────────────────────────────────────
        wb = openpyxl.Workbook()

        # Sheet 1 – Store Summary
        ws1 = wb.active
        ws1.title = "Store Summary"
        summary_headers = [
            "Store", "City",
            "Total Count", "Total Amount (INR)",
            "Approved Count", "Approved Amount (INR)",
            "Pending Count",  "Pending Amount (INR)",
            "Rejected Count",
        ]
        ws1.append(summary_headers)
        for s in store_summary.values():
            ws1.append([
                s["store_name"], s["city"],
                s["total_count"], s["total_amount"],
                s["approved_count"], s["approved_amount"],
                s["pending_count"],  s["pending_amount"],
                s["rejected_count"],
            ])

        # Sheet 2 – Detail
        ws2 = wb.create_sheet("Credit Notes Detail")
        if detail_rows:
            ws2.append(list(detail_rows[0].keys()))
            for row in detail_rows:
                ws2.append([str(v) if v is not None else "" for v in row.values()])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"credit_report_{start}_to_{end}"
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return resp


# ─────────────────────────────────────────────────────────────────────────────
# 4. Store-wise Revenue Time-series
# ─────────────────────────────────────────────────────────────────────────────

class StoreWiseRevenueReportView(SuperAdminMixin, APIView):
    """
    GET /api/superadmin/reports/store-wise/revenue/

    Always returns an Excel (.xlsx) download.
    Rows: Period | Store | Orders | Revenue (INR)

    Query params:
      period   : week | month | year | custom  (default: month)
      start_date, end_date – when period=custom
      group_by : day | week | month  (auto-selected if omitted)
      store_id : int  (optional)
    """

    def get(self, request):
        err = self._check(request)
        if err:
            return err

        from dreamspharmaapp.models import SalesOrder

        start, end = _resolve_period(request)
        period = request.query_params.get("period", "month").lower()
        store_filter = request.query_params.get("store_id")

        group_by = request.query_params.get("group_by")
        if not group_by:
            group_by = {"week": "day", "month": "week", "year": "month"}.get(period, "month")

        trunc = _trunc_fn(group_by)

        orders = SalesOrder.objects.filter(
            ord_date__gte=start,
            ord_date__lte=end,
        )
        if store_filter:
            orders = orders.filter(fulfilling_store_id=store_filter)

        trend = (
            orders
            .annotate(bucket=trunc("ord_date"))
            .values("bucket", "fulfilling_store__name")
            .annotate(
                order_count=Count("id"),
                revenue=Sum("order_total"),
                avg_order=Avg("order_total"),
            )
            .order_by("bucket", "fulfilling_store__name")
        )

        rows = []
        for row in trend:
            if not row["bucket"]:
                continue
            rows.append({
                "Period":        str(row["bucket"].date()) if hasattr(row["bucket"], "date") else str(row["bucket"]),
                "Store":         row["fulfilling_store__name"] or "",
                "Orders":        row["order_count"],
                "Revenue (INR)": float(row["revenue"] or 0),
                "Avg Order (INR)": round(float(row["avg_order"] or 0), 2),
            })

        filename = f"revenue_report_{start}_to_{end}_by_{group_by}"
        return _excel_response(rows, filename)


# ─────────────────────────────────────────────────────────────────────────────
# 5. Store-wise Retailer Activity Report
# ─────────────────────────────────────────────────────────────────────────────

class StoreWiseRetailerActivityReportView(SuperAdminMixin, APIView):
    """
    GET /api/superadmin/reports/store-wise/retailer-activity/

    Always returns an Excel (.xlsx) download.
    Per-retailer: orders, spend, avg order value, last order date,
    credit notes, and activity status.

    Query params:
      period     : week | month | year | custom  (default: month)
      start_date, end_date – when period=custom
      store_id   : int  (optional)
    """

    def get(self, request):
        err = self._check(request)
        if err:
            return err

        from dreamspharmaapp.models import SalesOrder, CreditNote
        from django.contrib.auth import get_user_model
        User = get_user_model()

        start, end = _resolve_period(request)
        store_filter = request.query_params.get("store_id")

        orders = SalesOrder.objects.filter(
            ord_date__gte=start,
            ord_date__lte=end,
        )
        if store_filter:
            orders = orders.filter(fulfilling_store_id=store_filter)

        # Aggregate by user_id
        activity = (
            orders
            .values(
                "user_id",
                "cust_name",
                "fulfilling_store__id",
                "fulfilling_store__name",
                "fulfilling_store__city",
            )
            .annotate(
                order_count=Count("id"),
                total_spent=Sum("order_total"),
                last_order_date=Max("ord_date"),
            )
            .order_by("-total_spent")
        )

        # Credit notes per user for the period
        cn_by_user: dict[str, dict] = {}
        cn_qs = CreditNote.objects.filter(
            created_at__date__gte=start,
            created_at__date__lte=end,
        )
        if store_filter:
            cn_qs = cn_qs.filter(store_id=store_filter)

        for cn in cn_qs.values("retailer_id").annotate(
            cn_count=Count("credit_note_id"),
            cn_amount=Sum("amount"),
        ):
            cn_by_user[str(cn["retailer_id"])] = {
                "cn_count":  cn["cn_count"],
                "cn_amount": float(cn["cn_amount"] or 0),
            }

        data = []
        for row in activity:
            uid = str(row["user_id"] or "")
            total_spent = float(row["total_spent"] or 0)
            order_count = row["order_count"]
            avg_val = round(total_spent / order_count, 2) if order_count else 0.0
            cn_info = cn_by_user.get(uid, {"cn_count": 0, "cn_amount": 0.0})

            data.append({
                "user_id":       uid,
                "retailer_name": row["cust_name"] or "—",
                "store":         row["fulfilling_store__name"] or "—",
                "city":          row["fulfilling_store__city"] or "—",
                "total_orders":  order_count,
                "total_spent":   round(total_spent, 2),
                "avg_order_value": avg_val,
                "last_order_date": str(row["last_order_date"]) if row["last_order_date"] else "—",
                "credit_notes_count":  cn_info["cn_count"],
                "credit_notes_amount": cn_info["cn_amount"],
                "status": "Active" if order_count >= 3 else ("Occasional" if order_count >= 1 else "Inactive"),
            })

        rows = [{
            "User ID":              r["user_id"],
            "Retailer":            r["retailer_name"],
            "Store":               r["store"],
            "City":                r["city"],
            "Total Orders":        r["total_orders"],
            "Total Spent (INR)":   r["total_spent"],
            "Avg Order (INR)":     r["avg_order_value"],
            "Last Order Date":     r["last_order_date"],
            "Credit Notes Count":  r["credit_notes_count"],
            "Credit Amount (INR)": r["credit_notes_amount"],
            "Status":              r["status"],
        } for r in data]

        filename = f"retailer_activity_{start}_to_{end}"
        return _excel_response(rows, filename)
